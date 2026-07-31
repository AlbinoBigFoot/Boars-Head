#!/usr/bin/env python3
"""Parse FactoryTalk display XMLs (batch 5) -> batch5.jsonl, then merge all batches to Excel."""
from __future__ import annotations

import html
import json
import re
import time
from collections import defaultdict
from pathlib import Path
from xml.etree import ElementTree as ET

DISPLAYS_DIR = Path(r"C:\Users\dylan.jones\Documents\Bors\Displays")
OUT_DIR = Path(r"C:\Users\dylan.jones\Documents\Bors\docs\ft-display-tags")

BATCH5_FILES = [
    "(STELLAR)Comp Trend.xml",
    "(STELLAR)Comp1.xml",
    "(STELLAR)Comp4.xml",
    "(STELLAR)Comp5.xml",
    "(STELLAR)Comp6.xml",
    "(STELLAR)Comp7.xml",
    "(STELLAR)CompTrendPOP.xml",
    "(STELLAR)Condensers.xml",
    "(STELLAR)CoolerTrend.xml",
    "(STELLAR)DockTrend.xml",
    "(STELLAR)Freezer Trend.xml",
    "(STELLAR)Machine Room Trend.xml",
    "(STELLAR)MachineRoom.xml",
    "(STELLAR)RDisks.xml",
    "(STELLAR)Runtimes.xml",
    "(STELLAR)SafetyRelay.xml",
    "pop_Security.xml",
]

FIELDS = [
    "display",
    "object_name",
    "link_base",
    "parameter",
    "parameter_description",
    "tag_expression",
    "resolved_tag",
    "plc_type_hint",
    "bh_component",
    "notes",
]

# Tag forms: {[RCP1]TAG}, {::[RCP1]Program:X.Y}, {/*S:0 {[RCP1]@Status}*/}
TAG_BRACE_RE = re.compile(
    r"\{(?:::+)?(?:/\*[^*]*\*/\s*)?\[([^\]]+)\]([^}]*)\}"
)
# Also bare {#102.Foo} placeholders (parameter refs) — keep as expression, not resolved PLC tags
PARAM_REF_RE = re.compile(r"\{#\d+(?:\.[^}]*)?\}")
ATTR_RE = re.compile(r'(\w+)="([^"]*)"')


def unescape(s: str) -> str:
    if not s:
        return ""
    return html.unescape(s.replace("&#xA;", "\n").replace("&quot;", '"'))


def extract_attrs(raw: str) -> dict:
    return {k: unescape(v) for k, v in ATTR_RE.findall(raw)}


def extract_plc_tags(expr: str) -> list[str]:
    """Return list of resolved PLC tag strings like {[RCP1]COMP[1]}."""
    if not expr:
        return []
    out = []
    for m in TAG_BRACE_RE.finditer(expr):
        shortcut = m.group(1)
        path = m.group(2).strip()
        # normalize :: prefix in original if present
        full = m.group(0)
        # Prefer canonical {[shortcut]path} without :: or /*S:0*/
        resolved = "{[%s]%s}" % (shortcut, path)
        out.append(resolved)
    return out


def plc_type_from_description(desc: str) -> str:
    if not desc:
        return ""
    d = desc
    # Prefer explicit (P_Xxx) patterns
    m = re.search(r"\b(P_[A-Za-z0-9]+(?:\s*,\s*P_[A-Za-z0-9]+)*)", d)
    if m:
        # take first type token
        first = m.group(1).split(",")[0].strip()
        return first
    if "Screw_Compressor" in d:
        return "Screw_Compressor"
    if re.search(r"\bVessel\b", d, re.I):
        return "Vessel"
    if "Runtime" in d:
        return "Runtime"
    if re.search(r"\bVSD\b|\bPF75[35]\b", d, re.I):
        return "P_VSD"
    if "Interlock" in d or "P_Intlk" in d:
        return "P_Intlk"
    if "Valve" in d:
        return "P_ValveSO"
    if "Motor" in d:
        return "P_Motor"
    return ""


def hint_from_tag(tag: str) -> str:
    t = tag.upper()
    if "COMP[" in t or re.search(r"COMP\d", t):
        return "Screw_Compressor"
    if "PUMP" in t:
        return "P_Motor"
    if "EEF" in t or "FAN" in t or "EXHAUST" in t:
        return "P_Motor"
    if "_CV" in t or "VALVE" in t or re.search(r"\bVL\b", t):
        return "P_ValveSO"
    if any(x in t for x in ("_PT", "SYS_PT", "_LT", "_TT", "_LVL", "LEVEL")):
        return "P_AIn"
    if any(x in t for x in ("HLCO", "LOLO", "HIHI", "OPL", "AD_", "MASTER_AD")):
        return "P_DIn"
    if "INTERLOCK" in t:
        return "P_Intlk"
    if "TOWER" in t or "CT[" in t or "COND" in t:
        return "CoolingTower"
    if "EVAP" in t or "CG_" in t or "CGDX" in t:
        return "Evaporator"
    if "HTR" in t or "HPR" in t or "LTR" in t or "VESSEL" in t or "ACCUM" in t:
        if "PUMP" in t:
            return "P_Motor"
        return "Vessel"
    return ""


def map_bh_component(
    plc_hint: str,
    tag: str,
    display: str,
    param_desc: str = "",
) -> str:
    d = display.upper()
    hint = (plc_hint or "").upper()
    t = (tag or "").upper()
    desc = (param_desc or "").upper()
    blob = " ".join([hint, t, desc, d])

    # Display-level overrides
    if d.startswith("(RA-BAS)") or "FACEPLATE" in d or d.endswith("-HELP") or d.endswith("-QUICK"):
        # still map device tags when present
        pass
    if "COMP" in d and "CONDENS" not in d:
        # Comp screens default compressor unless clearly otherwise
        if "COMP[" in t or "SCREW_COMPRESSOR" in hint or "COMP" in t:
            return "Compressor"

    if "SCREW_COMPRESSOR" in hint or "COMP[" in t:
        return "Compressor"
    if "P_VALVE" in hint or "VALVE" in hint:
        return "Valve"
    if "_CV" in t or "VALVE" in t:
        return "Valve"
    if "P_VSD" in hint or "PF75" in hint or "VSD" in hint:
        return "VFD"
    if "P_MOTOR" in hint or "MOTOR" in hint:
        if any(x in blob for x in ("FAN", "EEF", "EXHAUST")):
            return "ExhaustFan"
        return "Pump"
    if "PUMP" in t:
        return "Pump"
    if any(x in blob for x in ("FAN", "EEF", "EXHAUST")):
        return "ExhaustFan"
    if "COOLINGTOWER" in hint or "TOWER" in t or "CT[" in t:
        return "CoolingTower"
    if "CONDENS" in d:
        if "TOWER" in blob or "CT" in t or "FAN" in blob:
            return "CoolingTower"
        # condensers screen often cooling towers / other
        if "COMP" in t:
            return "Compressor"
        if "PUMP" in t:
            return "Pump"
        return "CoolingTower" if ("COND" in t or "CT" in t) else "Other"
    if "EVAP" in hint or "CG_" in t or "CGDX" in t or "EVAP" in t:
        return "Evaporator"
    if "VESSEL" in hint or any(x in t for x in ("HTR", "HPR", "LTR", "ACCUM")):
        # level switches on vessels -> Tank
        if "P_DIN" in hint or any(x in t for x in ("HLCO", "LOLO", "HIHI", "OPL")):
            return "Tank"
        if "P_AIN" in hint or "_LVL" in t or "LT_" in t:
            return "Tank"
        if "#110" in desc or "VESSEL NAME" in desc:
            return "Tank"
        return "Tank"
    if "P_AIN" in hint or "P_AOUT" in hint or "P_AICHAN" in hint:
        return "Sensor"
    if "P_DIN" in hint or "P_DOUT" in hint:
        if any(x in t for x in ("HLCO", "LOLO", "HIHI", "OPL", "LEVEL")):
            return "Tank"
        return "Sensor"
    if "P_INTLK" in hint or "INTERLOCK" in t:
        return "Other"
    if "P_ALARM" in hint or "ALARM" in t:
        return "Other"
    if "RUNTIME" in hint or "RUNTIME" in t:
        return "Other"
    if "P_PIDE" in hint:
        return "Analog"
    if not tag and (d.startswith("(RA-BAS)") or "FACEPLATE" in d):
        return "Faceplate"
    if tag:
        return "Other"
    return "Unknown"


def nearest_object_name(elem: ET.Element, parents: list[ET.Element]) -> str:
    # Walk up for name=
    for e in [elem] + list(reversed(parents)):
        name = e.attrib.get("name")
        if name:
            return name
    return ""


def parse_display(path: Path) -> list[dict]:
    display = path.stem
    text = path.read_text(encoding="utf-8", errors="replace")
    rows: list[dict] = []
    seen: set[tuple] = set()

    def add_row(
        object_name="",
        link_base="",
        parameter="",
        parameter_description="",
        tag_expression="",
        resolved_tag="",
        plc_type_hint="",
        bh_component="",
        notes="",
    ):
        if not plc_type_hint:
            plc_type_hint = plc_type_from_description(parameter_description) or hint_from_tag(
                resolved_tag or tag_expression
            )
        if not bh_component:
            bh_component = map_bh_component(
                plc_type_hint, resolved_tag, display, parameter_description
            )
        # Display-level Comp* override
        if display.startswith("(STELLAR)Comp") and "Condens" not in display:
            if resolved_tag and ("COMP[" in resolved_tag.upper() or "COMP" in display):
                if bh_component in ("Unknown", "Other", "Faceplate") or "COMP[" in resolved_tag.upper():
                    if "COMP[" in resolved_tag.upper() or "SCREW" in (plc_type_hint or "").upper():
                        bh_component = "Compressor"
        key = (
            display,
            object_name,
            link_base,
            parameter,
            resolved_tag or tag_expression,
            bh_component,
        )
        if key in seen:
            return
        seen.add(key)
        rows.append(
            {
                "display": display,
                "object_name": object_name,
                "link_base": link_base,
                "parameter": parameter,
                "parameter_description": parameter_description,
                "tag_expression": tag_expression,
                "resolved_tag": resolved_tag,
                "plc_type_hint": plc_type_hint,
                "bh_component": bh_component,
                "notes": notes,
            }
        )

    # --- Pass 1: regex for parameters (reliable on large files) ---
    # Find parameter elements with surrounding context for object/linkBase
    # Split by groups roughly using iterative scan
    for m in re.finditer(r"<parameter\s+([^>]*?)\s*/>", text):
        attrs = extract_attrs(m.group(1))
        pname = attrs.get("name", "")
        pdesc = attrs.get("description", "")
        pval = attrs.get("value", "")
        # Look back for nearest name= and linkBaseObject=
        start = max(0, m.start() - 2500)
        window = text[start : m.start()]
        obj_names = re.findall(r'\bname="([^"]+)"', window)
        link_bases = re.findall(r'linkBaseObject="([^"]+)"', window)
        object_name = obj_names[-1] if obj_names else ""
        link_base = unescape(link_bases[-1]) if link_bases else ""

        tags = extract_plc_tags(pval)
        plc_hint = plc_type_from_description(pdesc)

        if tags:
            for tag in tags:
                add_row(
                    object_name=object_name,
                    link_base=link_base,
                    parameter=pname,
                    parameter_description=pdesc,
                    tag_expression=pval,
                    resolved_tag=tag,
                    plc_type_hint=plc_hint or hint_from_tag(tag),
                    notes="from <parameter>",
                )
        else:
            # Non-tag parameters still useful (#110 vessel names, display names)
            if pname in ("#102", "#110", "#111", "#124", "#101") or pval or pdesc:
                notes = "parameter value (non-PLC or string)"
                if pname == "#110":
                    notes = "vessel name parameter"
                add_row(
                    object_name=object_name,
                    link_base=link_base,
                    parameter=pname,
                    parameter_description=pdesc,
                    tag_expression=pval,
                    resolved_tag="",
                    plc_type_hint=plc_hint,
                    bh_component="Tank" if pname == "#110" else map_bh_component(plc_hint, "", display, pdesc),
                    notes=notes,
                )

    # --- Pass 2: connection expressions and any expression= with PLC tags ---
    for m in re.finditer(r'\bexpression="([^"]*)"', text):
        expr = unescape(m.group(1))
        if not expr or expr.strip() == "":
            continue
        tags = extract_plc_tags(expr)
        if not tags:
            continue
        start = max(0, m.start() - 2000)
        window = text[start : m.start()]
        obj_names = re.findall(r'\bname="([^"]+)"', window)
        link_bases = re.findall(r'linkBaseObject="([^"]+)"', window)
        object_name = obj_names[-1] if obj_names else ""
        link_base = unescape(link_bases[-1]) if link_bases else ""
        for tag in tags:
            add_row(
                object_name=object_name,
                link_base=link_base,
                parameter="",
                parameter_description="",
                tag_expression=expr[:500],
                resolved_tag=tag,
                plc_type_hint=hint_from_tag(tag),
                notes="from expression/connection",
            )

    # --- Pass 3: captions / tooltips / pressAction that embed {[...]} ---
    for attr in ("caption", "toolTipText", "pressAction", "releaseAction", "value"):
        for m in re.finditer(rf'\b{attr}="([^"]*)"', text):
            expr = unescape(m.group(1))
            tags = extract_plc_tags(expr)
            if not tags:
                continue
            start = max(0, m.start() - 1500)
            window = text[start : m.start()]
            obj_names = re.findall(r'\bname="([^"]+)"', window)
            link_bases = re.findall(r'linkBaseObject="([^"]+)"', window)
            object_name = obj_names[-1] if obj_names else ""
            link_base = unescape(link_bases[-1]) if link_bases else ""
            for tag in tags:
                add_row(
                    object_name=object_name,
                    link_base=link_base,
                    parameter="",
                    parameter_description="",
                    tag_expression=expr[:500],
                    resolved_tag=tag,
                    plc_type_hint=hint_from_tag(tag),
                    notes=f"from {attr}",
                )

    # --- Pass 4: XML tree for structured params under groups (backup) ---
    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        add_row(notes=f"XML parse warning: {e}")
        return rows

    def walk(elem: ET.Element, parents: list[ET.Element]):
        if elem.tag == "parameter":
            pname = elem.attrib.get("name", "")
            pdesc = unescape(elem.attrib.get("description", ""))
            pval = unescape(elem.attrib.get("value", ""))
            object_name = nearest_object_name(elem, parents)
            link_base = ""
            for e in [elem] + list(reversed(parents)):
                if e.attrib.get("linkBaseObject"):
                    link_base = unescape(e.attrib["linkBaseObject"])
                    break
            tags = extract_plc_tags(pval)
            plc_hint = plc_type_from_description(pdesc)
            if tags:
                for tag in tags:
                    add_row(
                        object_name=object_name,
                        link_base=link_base,
                        parameter=pname,
                        parameter_description=pdesc,
                        tag_expression=pval,
                        resolved_tag=tag,
                        plc_type_hint=plc_hint or hint_from_tag(tag),
                        notes="from XML tree parameter",
                    )
            elif pname in ("#102", "#110", "#111", "#124"):
                add_row(
                    object_name=object_name,
                    link_base=link_base,
                    parameter=pname,
                    parameter_description=pdesc,
                    tag_expression=pval,
                    resolved_tag="",
                    plc_type_hint=plc_hint,
                    bh_component="Tank" if pname == "#110" else "",
                    notes="from XML tree parameter (non-PLC)",
                )
        # connections
        if elem.tag == "connection":
            expr = unescape(elem.attrib.get("expression", ""))
            tags = extract_plc_tags(expr)
            if tags:
                object_name = nearest_object_name(elem, parents)
                link_base = ""
                for e in list(reversed(parents)):
                    if e.attrib.get("linkBaseObject"):
                        link_base = unescape(e.attrib["linkBaseObject"])
                        break
                for tag in tags:
                    add_row(
                        object_name=object_name,
                        link_base=link_base,
                        tag_expression=expr[:500],
                        resolved_tag=tag,
                        plc_type_hint=hint_from_tag(tag),
                        notes="from connection element",
                    )
        for child in list(elem):
            walk(child, parents + [elem])

    walk(root, [])
    return rows


def write_jsonl(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as f:
        for r in rows:
            obj = {k: r.get(k, "") for k in FIELDS}
            f.write(json.dumps(obj, ensure_ascii=False) + "\n")


def load_jsonl(path: Path) -> list[dict]:
    rows = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            rows.append(json.loads(line))
    return rows


def dedupe_rows(rows: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for r in rows:
        key = (
            r.get("display", ""),
            r.get("object_name", ""),
            r.get("link_base", ""),
            r.get("parameter", ""),
            r.get("tag_expression", ""),
            r.get("resolved_tag", ""),
            r.get("bh_component", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def wait_for_batches(timeout_s: int = 600) -> dict[str, Path]:
    needed = [f"batch{i}.jsonl" for i in range(1, 6)]
    deadline = time.time() + timeout_s
    found = {}
    while time.time() < deadline:
        for name in needed:
            p = OUT_DIR / name
            if name not in found and p.exists() and p.stat().st_size > 0:
                # ensure file not still being written: size stable
                sz1 = p.stat().st_size
                time.sleep(0.5)
                sz2 = p.stat().st_size
                if sz1 == sz2:
                    found[name] = p
        if len(found) == 5:
            return found
        missing = [n for n in needed if n not in found]
        print(f"Waiting for batches: missing {missing} ({len(found)}/5)...")
        time.sleep(5)
    return found


def sample_tags(tags, n=5):
    uniq = []
    for t in tags:
        if t and t not in uniq:
            uniq.append(t)
        if len(uniq) >= n:
            break
    return "; ".join(uniq)


def build_excel(all_rows: list[dict], out_path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def write_sheet(ws, headers, data_rows):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in data_rows:
            ws.append([row.get(h, "") if isinstance(row, dict) else row[i] for i, h in enumerate(headers)] if isinstance(row, dict) else list(row))
        for i, h in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, len(h) + 2))

    # 1. All_Bindings
    ws = wb.active
    ws.title = "All_Bindings"
    write_sheet(ws, FIELDS, all_rows)

    # 2. By_Display
    by_disp = defaultdict(lambda: {"tags": set(), "comps": set(), "samples": []})
    for r in all_rows:
        d = r.get("display", "")
        tag = r.get("resolved_tag") or ""
        if tag:
            by_disp[d]["tags"].add(tag)
            if len(by_disp[d]["samples"]) < 8:
                by_disp[d]["samples"].append(tag)
        comp = r.get("bh_component") or ""
        if comp and comp not in ("Unknown",):
            by_disp[d]["comps"].add(comp)
    ws2 = wb.create_sheet("By_Display")
    rows2 = []
    for d in sorted(by_disp.keys()):
        info = by_disp[d]
        rows2.append(
            {
                "display": d,
                "unique_tag_count": len(info["tags"]),
                "components_found": ", ".join(sorted(info["comps"])),
                "sample_tags": sample_tags(info["samples"], 5),
            }
        )
    write_sheet(ws2, ["display", "unique_tag_count", "components_found", "sample_tags"], rows2)

    # 3. By_Component
    by_comp = defaultdict(lambda: {"tags": set(), "displays": set(), "samples": []})
    for r in all_rows:
        c = r.get("bh_component") or "Unknown"
        tag = r.get("resolved_tag") or ""
        d = r.get("display") or ""
        by_comp[c]["displays"].add(d)
        if tag:
            by_comp[c]["tags"].add(tag)
            if len(by_comp[c]["samples"]) < 8:
                by_comp[c]["samples"].append(tag)
    ws3 = wb.create_sheet("By_Component")
    rows3 = []
    for c in sorted(by_comp.keys()):
        info = by_comp[c]
        rows3.append(
            {
                "bh_component": c,
                "tag_count": len(info["tags"]),
                "displays": "; ".join(sorted(info["displays"])),
                "sample_tags": sample_tags(info["samples"], 5),
            }
        )
    write_sheet(ws3, ["bh_component", "tag_count", "displays", "sample_tags"], rows3)

    # 4. Unique_Tags
    by_tag = defaultdict(lambda: {"hints": set(), "comps": set(), "displays": set()})
    for r in all_rows:
        tag = r.get("resolved_tag") or ""
        if not tag:
            continue
        by_tag[tag]["displays"].add(r.get("display") or "")
        if r.get("plc_type_hint"):
            by_tag[tag]["hints"].add(r["plc_type_hint"])
        if r.get("bh_component"):
            by_tag[tag]["comps"].add(r["bh_component"])
    ws4 = wb.create_sheet("Unique_Tags")
    rows4 = []
    for tag in sorted(by_tag.keys()):
        info = by_tag[tag]
        # prefer non-empty specific hint/comp
        hint = sorted(info["hints"], key=lambda x: (0 if x else 1, x))[0] if info["hints"] else ""
        comps = [c for c in info["comps"] if c not in ("Unknown", "")]
        comp = sorted(comps)[0] if comps else (sorted(info["comps"])[0] if info["comps"] else "")
        rows4.append(
            {
                "resolved_tag": tag,
                "plc_type_hint": hint,
                "bh_component": comp,
                "displays": "; ".join(sorted(info["displays"])),
            }
        )
    write_sheet(ws4, ["resolved_tag", "plc_type_hint", "bh_component", "displays"], rows4)

    # 5. MachineRoom
    ws5 = wb.create_sheet("MachineRoom")
    mr_rows = [r for r in all_rows if "MachineRoom" in (r.get("display") or "")]
    write_sheet(ws5, FIELDS, mr_rows)

    # 6. STELLAR_Screens
    ws6 = wb.create_sheet("STELLAR_Screens")
    st_rows = [
        r
        for r in all_rows
        if (r.get("display") or "").startswith("(STELLAR)")
    ]
    write_sheet(ws6, FIELDS, st_rows)

    # 7. RA_BAS_Faceplates
    ws7 = wb.create_sheet("RA_BAS_Faceplates")
    ra_rows = [
        r
        for r in all_rows
        if (r.get("display") or "").startswith("(RA-BAS)")
    ]
    write_sheet(ws7, FIELDS, ra_rows)

    # 8. README
    ws8 = wb.create_sheet("README")
    readme = [
        ["FT Display Tag Map"],
        [""],
        ["How generated"],
        [
            "Parsed FactoryTalk View SE display XML exports under Displays/ "
            "into batch1–batch5 JSONL, then merged with openpyxl."
        ],
        [""],
        ["Field definitions"],
        ["display", "Display filename without .xml"],
        ["object_name", "Nearest named object/group containing the binding"],
        ["link_base", "linkBaseObject for global/reference objects"],
        ["parameter", "FactoryTalk parameter placeholder (e.g. #102, #110)"],
        ["parameter_description", "Parameter description attribute (often includes PLC UDT type)"],
        ["tag_expression", "Raw expression or parameter value"],
        ["resolved_tag", "Extracted {[shortcut]tag} when present"],
        ["plc_type_hint", "Inferred PLC AOI/UDT type (P_Motor, Screw_Compressor, P_ValveSO, ...)"],
        ["bh_component", "Mapped BH HMI component family"],
        ["notes", "Extraction source / caveats"],
        [""],
        ["Limitations"],
        [
            "Global objects referenced via linkBaseObject may embed additional tags only in library GFX/XML; "
            "those library definitions are not fully expanded here."
        ],
        [
            "Expressions using {#102.Member} without a local #102 value resolve only at runtime from the "
            "parameter substitution on the referencing display."
        ],
        [
            "ActiveX / embedded browser content and encrypted VBA are not parsed for tags."
        ],
        [
            "bh_component mapping is heuristic from plc_type_hint + tag name + display name."
        ],
        [""],
        ["Sheets"],
        ["All_Bindings", "Every binding row from all batches (near-duplicate deduped)"],
        ["By_Display", "Per-display unique tag counts and components"],
        ["By_Component", "Per bh_component tag counts and displays"],
        ["Unique_Tags", "One row per resolved_tag with displays list"],
        ["MachineRoom", "Rows where display contains MachineRoom"],
        ["STELLAR_Screens", "(STELLAR)* process screens only"],
        ["RA_BAS_Faceplates", "(RA-BAS)* faceplates/help/quick only"],
    ]
    for row in readme:
        ws8.append(row)
    ws8.column_dimensions["A"].width = 28
    ws8.column_dimensions["B"].width = 100

    wb.save(out_path)


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    all_batch5: list[dict] = []
    print("=== Parsing batch 5 ===")
    for fname in BATCH5_FILES:
        path = DISPLAYS_DIR / fname
        if not path.exists():
            print(f"MISSING: {fname}")
            continue
        rows = parse_display(path)
        print(f"  {fname}: {len(rows)} rows")
        all_batch5.extend(rows)

    all_batch5 = dedupe_rows(all_batch5)
    out5 = OUT_DIR / "batch5.jsonl"
    write_jsonl(out5, all_batch5)
    print(f"Wrote {len(all_batch5)} rows -> {out5}")

    # MachineRoom sanity
    mr = [r for r in all_batch5 if r["display"] == "(STELLAR)MachineRoom"]
    mr_tags = sorted({r["resolved_tag"] for r in mr if r.get("resolved_tag")})
    print(f"MachineRoom rows={len(mr)} unique_resolved_tags={len(mr_tags)}")
    p102 = [r for r in mr if r.get("parameter") == "#102"]
    p110 = [r for r in mr if r.get("parameter") == "#110"]
    print(f"  #102 params={len(p102)}  #110 params={len(p110)}")

    print("=== Waiting for batch1-4 ===")
    found = wait_for_batches(600)
    print("Found:", sorted(found.keys()))
    if len(found) < 5:
        print("WARNING: incomplete batches; merging what is available")

    merged: list[dict] = []
    for i in range(1, 6):
        name = f"batch{i}.jsonl"
        p = OUT_DIR / name
        if p.exists():
            rows = load_jsonl(p)
            print(f"  loaded {name}: {len(rows)}")
            merged.extend(rows)
        else:
            print(f"  MISSING {name}")

    merged = dedupe_rows(merged)
    xlsx = OUT_DIR / "FT_Display_Tag_Map.xlsx"
    build_excel(merged, xlsx)
    print(f"Excel: {xlsx}")
    print(f"Total rows: {len(merged)}")
    print(
        "Sheets: All_Bindings, By_Display, By_Component, Unique_Tags, "
        "MachineRoom, STELLAR_Screens, RA_BAS_Faceplates, README"
    )


if __name__ == "__main__":
    main()
