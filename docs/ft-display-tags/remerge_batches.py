#!/usr/bin/env python3
"""Re-merge batch1–batch5 JSONL into FT_Display_Tag_Map.xlsx (after batch2 completed)."""
from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

OUT_DIR = Path(r"C:\Users\dylan.jones\Documents\Bors\docs\ft-display-tags")
XLSX = OUT_DIR / "FT_Display_Tag_Map.xlsx"

FIELDS = [
    "display",
    "object_name",
    "link_base",
    "parameter",
    "parameter_description",
    "tag_expression",
    "resolved_tag",
    "plc_type_hint",
    "bh_component",
    "notes",
]


def load_jsonl(path: Path) -> list[dict]:
    rows = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line:
                rows.append(json.loads(line))
    return rows


def sample_tags(tags, n=5):
    uniq = []
    for t in tags:
        if t and t not in uniq:
            uniq.append(t)
        if len(uniq) >= n:
            break
    return "; ".join(uniq)


def write_sheet(ws, headers, data_rows):
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in data_rows:
        if isinstance(row, dict):
            ws.append([row.get(h, "") for h in headers])
        else:
            ws.append(list(row))
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(12, len(h) + 2))


def dedupe_rows(rows: list[dict]) -> list[dict]:
    """Near-identical: same display + tag_expression + parameter + object_name."""
    seen = set()
    out = []
    for r in rows:
        key = (
            r.get("display", ""),
            r.get("tag_expression", ""),
            r.get("parameter", ""),
            r.get("object_name", ""),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(r)
    return out


def build_excel(all_rows: list[dict], out_path: Path) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "All_Bindings"
    write_sheet(ws, FIELDS, all_rows)

    by_disp = defaultdict(lambda: {"tags": set(), "comps": set(), "samples": []})
    for r in all_rows:
        d = r.get("display", "")
        tag = r.get("resolved_tag") or ""
        if tag:
            by_disp[d]["tags"].add(tag)
            if len(by_disp[d]["samples"]) < 8:
                by_disp[d]["samples"].append(tag)
        comp = r.get("bh_component") or ""
        if comp and comp not in ("Unknown",):
            by_disp[d]["comps"].add(comp)
    ws2 = wb.create_sheet("By_Display")
    rows2 = []
    for d in sorted(by_disp.keys()):
        info = by_disp[d]
        rows2.append(
            {
                "display": d,
                "unique_tag_count": len(info["tags"]),
                "components_found": ", ".join(sorted(info["comps"])),
                "sample_tags": sample_tags(info["samples"], 5),
            }
        )
    write_sheet(ws2, ["display", "unique_tag_count", "components_found", "sample_tags"], rows2)

    by_comp = defaultdict(lambda: {"tags": set(), "displays": set(), "samples": []})
    for r in all_rows:
        c = r.get("bh_component") or "Unknown"
        tag = r.get("resolved_tag") or ""
        d = r.get("display") or ""
        by_comp[c]["displays"].add(d)
        if tag:
            by_comp[c]["tags"].add(tag)
            if len(by_comp[c]["samples"]) < 8:
                by_comp[c]["samples"].append(tag)
    ws3 = wb.create_sheet("By_Component")
    rows3 = []
    for c in sorted(by_comp.keys()):
        info = by_comp[c]
        rows3.append(
            {
                "bh_component": c,
                "tag_count": len(info["tags"]),
                "displays": "; ".join(sorted(info["displays"])),
                "sample_tags": sample_tags(info["samples"], 5),
            }
        )
    write_sheet(ws3, ["bh_component", "tag_count", "displays", "sample_tags"], rows3)

    by_tag = defaultdict(lambda: {"hints": set(), "comps": set(), "displays": set()})
    for r in all_rows:
        tag = r.get("resolved_tag") or ""
        if not tag:
            continue
        by_tag[tag]["displays"].add(r.get("display") or "")
        if r.get("plc_type_hint"):
            by_tag[tag]["hints"].add(r["plc_type_hint"])
        if r.get("bh_component"):
            by_tag[tag]["comps"].add(r["bh_component"])
    ws4 = wb.create_sheet("Unique_Tags")
    rows4 = []
    for tag in sorted(by_tag.keys()):
        info = by_tag[tag]
        hint = (
            sorted(info["hints"], key=lambda x: (0 if x else 1, x))[0]
            if info["hints"]
            else ""
        )
        comps = [c for c in info["comps"] if c not in ("Unknown", "")]
        comp = (
            sorted(comps)[0]
            if comps
            else (sorted(info["comps"])[0] if info["comps"] else "")
        )
        rows4.append(
            {
                "resolved_tag": tag,
                "plc_type_hint": hint,
                "bh_component": comp,
                "displays": "; ".join(sorted(info["displays"])),
            }
        )
    write_sheet(ws4, ["resolved_tag", "plc_type_hint", "bh_component", "displays"], rows4)

    ws5 = wb.create_sheet("MachineRoom")
    write_sheet(
        ws5,
        FIELDS,
        [r for r in all_rows if "MachineRoom" in (r.get("display") or "")],
    )

    ws6 = wb.create_sheet("STELLAR_Screens")
    write_sheet(
        ws6,
        FIELDS,
        [r for r in all_rows if (r.get("display") or "").startswith("(STELLAR)")],
    )

    ws7 = wb.create_sheet("RA_BAS_Faceplates")
    write_sheet(
        ws7,
        FIELDS,
        [r for r in all_rows if (r.get("display") or "").startswith("(RA-BAS)")],
    )

    ws8 = wb.create_sheet("README")
    readme = [
        ["FT Display Tag Map"],
        [""],
        ["Regenerated"],
        [
            "Workbook regenerated after batch2.jsonl completed (~3602 rows). "
            "Merged batch1–batch5."
        ],
        [""],
        ["How generated"],
        [
            "Parsed FactoryTalk View SE display XML exports under Displays/ "
            "into batch1–batch5 JSONL, then merged with openpyxl."
        ],
        [""],
        ["Field definitions"],
        ["display", "Display filename without .xml"],
        ["object_name", "Nearest named object/group containing the binding"],
        ["link_base", "linkBaseObject for global/reference objects"],
        ["parameter", "FactoryTalk parameter placeholder (e.g. #102, #110)"],
        [
            "parameter_description",
            "Parameter description attribute (often includes PLC UDT type)",
        ],
        ["tag_expression", "Raw expression or parameter value"],
        [
            "resolved_tag",
            "Extracted {[shortcut]tag} when present; faceplates often have empty resolved_tag",
        ],
        [
            "plc_type_hint",
            "Inferred PLC AOI/UDT type (P_Motor, Screw_Compressor, P_ValveSO, ...)",
        ],
        ["bh_component", "Mapped BH HMI component family"],
        ["notes", "Extraction source / caveats"],
        [""],
        ["Limitations"],
        [
            "Faceplates (RA-BAS) often have empty resolved_tag — parameters/expressions "
            "resolve at runtime via #nnn placeholders."
        ],
        [
            "Global objects referenced via linkBaseObject may embed additional tags only "
            "in library GFX/XML; those library definitions are not fully expanded here."
        ],
        [
            "Expressions using {#102.Member} without a local #102 value resolve only at "
            "runtime from the parameter substitution on the referencing display."
        ],
        ["ActiveX / embedded browser content and encrypted VBA are not parsed for tags."],
        [
            "bh_component mapping is heuristic from plc_type_hint + tag name + display name."
        ],
        [""],
        ["Dedupe"],
        [
            "Near-identical rows removed when display + tag_expression + parameter + "
            "object_name match."
        ],
        [""],
        ["Sheets"],
        ["All_Bindings", "Every binding row from all batches (near-duplicate deduped)"],
        ["By_Display", "Per-display unique tag counts and components"],
        ["By_Component", "Per bh_component tag counts and displays"],
        ["Unique_Tags", "One row per resolved_tag with displays list"],
        ["MachineRoom", "Rows where display contains MachineRoom"],
        ["STELLAR_Screens", "(STELLAR)* process screens only (not RA-BAS)"],
        ["RA_BAS_Faceplates", "(RA-BAS)* faceplates/help/quick only"],
    ]
    for row in readme:
        ws8.append(row)
    ws8.column_dimensions["A"].width = 28
    ws8.column_dimensions["B"].width = 100

    wb.save(out_path)
    print(f"Sheets: {wb.sheetnames}")


def main():
    counts = {}
    merged: list[dict] = []
    for i in range(1, 6):
        p = OUT_DIR / f"batch{i}.jsonl"
        if not p.exists() or p.stat().st_size == 0:
            raise SystemExit(f"MISSING or empty: {p}")
        rows = load_jsonl(p)
        counts[p.name] = len(rows)
        merged.extend(rows)
        print(f"{p.name}: {len(rows)}")

    before = len(merged)
    print(f"Total before dedupe: {before}")
    deduped = dedupe_rows(merged)
    after = len(deduped)
    print(f"Total after dedupe: {after} (removed {before - after})")

    build_excel(deduped, XLSX)
    print(f"Excel: {XLSX}")
    print(f"exists={XLSX.exists()} size={XLSX.stat().st_size}")


if __name__ == "__main__":
    main()
